"""
Script to create Excel file with Test Cases organized by subtype
Each subtype has headers and is separated by blank rows
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# Black-Box sheet
bb_sheet = wb.active
bb_sheet.title = "Black-Box Test Cases"

# White-Box sheet
wb_sheet = wb.create_sheet("White-Box Test Cases")

# Styling
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
subtype_font = Font(bold=True, size=12)
subtype_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Column headers for Black-Box
bb_headers = [
    "Test_Case_ID",
    "Technique",
    "Module",
    "Input_Field",
    "Boundary_Checked",
    "Test_Value",
    "Expected_Result",
    "Request_Method",
    "Request_URL",
    "Request_Body",
    "DB_Preconditions"
]

# Column headers for White-Box
wb_headers = [
    "Test_Case_ID",
    "Technique",
    "Module",
    "Function",
    "Inputs_or_Setup",
    "Code_Lines_Executed_(Description)",
    "Expected_Output",
    "Request_Method",
    "Request_URL",
    "Request_Body",
    "DB_Preconditions"
]

# BLACK-BOX DATA
# A. Equivalence Class Partitioning
ecp_data = [
    ["TC-BB-ECP-001", "Equivalence Class Partitioning", "Admin Authentication", "email, password", "No", 
     'email: "admin@healthconnect.com", password: "#1ap@NITK"', 
     "Status 200, admin login successful, JWT token set in cookie, admin data returned", 
     "POST", "http://localhost:5000/api/auth/admin/login", 
     '{"email": "admin@healthconnect.com", "password": "#1ap@NITK"}', 
     'Admin account exists with email "admin@healthconnect.com" and isActive=true'],
    
    ["TC-BB-ECP-002", "Equivalence Class Partitioning", "Admin Authentication", "email, password", "No", 
     'email: "nonexistent@admin.com", password: "SomePassword123"', 
     'Status 401, error message: "Invalid credentials"', 
     "POST", "http://localhost:5000/api/auth/admin/login", 
     '{"email": "nonexistent@admin.com", "password": "SomePassword123"}', 
     'No admin account exists with email "nonexistent@admin.com"'],
    
    ["TC-BB-ECP-003", "Equivalence Class Partitioning", "Doctor Registration", 
     "name, email, password, role, specialization, experience, qualification", "No", 
     'name: "Dr. Sarah Johnson", email: "sarah.j@hospital.com", password: "SecurePass123!", role: "doctor", specialization: "Cardiology", experience: 8, qualification: "MD, MBBS"', 
     "Status 201, registration successful, OTP sent to email, user created with isEmailVerified=false", 
     "POST", "http://localhost:5000/api/auth/register", 
     '{"name": "Dr. Sarah Johnson", "email": "sarah.j@hospital.com", "password": "SecurePass123!", "role": "doctor", "specialization": "Cardiology", "experience": 8, "qualification": "MD, MBBS"}', 
     'No existing doctor or patient with email "sarah.j@hospital.com"'],
    
    ["TC-BB-ECP-004", "Equivalence Class Partitioning", "Patient Registration", "role", "No", 
     'name: "John Doe", email: "john.doe@email.com", password: "Pass1234!", role: "admin" (invalid role for registration)', 
     'Status 400, validation error for invalid role (only "doctor" or "patient" allowed in registration)', 
     "POST", "http://localhost:5000/api/auth/register", 
     '{"name": "John Doe", "email": "john.doe@email.com", "password": "Pass1234!", "role": "admin"}', 
     "None"],
    
    ["TC-BB-ECP-005", "Equivalence Class Partitioning", "Appointment Management", 
     "doctorId, patientId, date, time, reason", "No", 
     'doctorId: "507f1f77bcf86cd799439011", patientId: "507f1f77bcf86cd799439012", date: "2025-11-05", time: "10:00 AM", reason: "Regular checkup"', 
     "Status 201, appointment created successfully, appointment ID returned", 
     "POST", "http://localhost:5000/api/appointments", 
     '{"doctorId": "507f1f77bcf86cd799439011", "patientId": "507f1f77bcf86cd799439012", "date": "2025-11-05", "time": "10:00 AM", "reason": "Regular checkup"}', 
     "Doctor with ID exists and isActive=true, Patient with ID exists and isEmailVerified=true, Time slot is available"]
]

# B. Boundary Value Analysis
bva_data = [
    ["TC-BB-BVA-001", "Boundary Value Analysis", "OTP Verification", "otp", "Yes - Minimum 6-digit OTP value", 
     'email: "patient@test.com", otp: "000000", role: "patient", purpose: "registration"', 
     "Status 400, OTP verification fails (invalid OTP - unless this is the actual OTP generated)", 
     "POST", "http://localhost:5000/api/auth/verify-otp", 
     '{"email": "patient@test.com", "otp": "000000", "role": "patient", "purpose": "registration"}', 
     'OTP exists for email "patient@test.com" with purpose "registration", OTP value is different from "000000"'],
    
    ["TC-BB-BVA-002", "Boundary Value Analysis", "OTP Verification", "otp", "Yes - Maximum 6-digit OTP value", 
     'email: "doctor@test.com", otp: "999999", role: "doctor", purpose: "login"', 
     "Status 400, OTP verification fails (invalid OTP - unless this is the actual OTP generated)", 
     "POST", "http://localhost:5000/api/auth/verify-otp", 
     '{"email": "doctor@test.com", "otp": "999999", "role": "doctor", "purpose": "login"}', 
     'OTP exists for email "doctor@test.com" with purpose "login", OTP value is different from "999999"'],
    
    ["TC-BB-BVA-003", "Boundary Value Analysis", "Doctor Registration", "experience", "Yes - Minimum experience value (0)", 
     'name: "Dr. Fresh Graduate", email: "newdoc@hospital.com", password: "Pass1234!", role: "doctor", specialization: "General Medicine", experience: 0, qualification: "MBBS"', 
     "Status 201, registration successful with experience=0, OTP sent", 
     "POST", "http://localhost:5000/api/auth/register", 
     '{"name": "Dr. Fresh Graduate", "email": "newdoc@hospital.com", "password": "Pass1234!", "role": "doctor", "specialization": "General Medicine", "experience": 0, "qualification": "MBBS"}', 
     'No existing user with email "newdoc@hospital.com"'],
    
    ["TC-BB-BVA-004", "Boundary Value Analysis", "Doctor Profile Update", "experience", "Yes - High experience value (50)", 
     "experience: 50", 
     "Status 200, profile updated successfully with experience=50", 
     "PUT", "http://localhost:5000/api/doctors/profile", 
     '{"experience": 50}', 
     "Authenticated doctor user exists, valid JWT token in cookie"],
    
    ["TC-BB-BVA-005", "Boundary Value Analysis", "Password Reset", "otp (time boundary)", "Yes - OTP expiry time (10 minutes)", 
     'email: "patient@test.com", otp: "123456" (created exactly 10 minutes ago), password: "NewPass123!", confirmPassword: "NewPass123!", role: "patient"', 
     "Status 400, OTP expired error message", 
     "POST", "http://localhost:5000/api/auth/reset-password", 
     '{"email": "patient@test.com", "otp": "123456", "password": "NewPass123!", "confirmPassword": "NewPass123!", "role": "patient"}', 
     'OTP exists with email "patient@test.com", expiresAt timestamp is exactly 10 minutes old']
]

# WHITE-BOX DATA
# A. Statement Coverage
sc_data = [
    ["TC-WB-SC-001", "Statement Coverage", "Authentication", "register() in authController.js", 
     'POST request with: name="Test User", email="new@test.com", password="Pass123!", role="patient"', 
     "Lines creating user with `isEmailVerified: false`, OTP generation via `OTP.createOTP()`, email sending via `sendOTPEmail()`, logging via `logAuthEvent()`, response with `requiresVerification: true`", 
     'User created in DB with isEmailVerified=false, OTP record created, email sent, response: {"success": true, "message": "Registration successful! Please check your email...", "data": {"email": "new@test.com", "requiresVerification": true}}', 
     "POST", "http://localhost:5000/api/auth/register", 
     '{"name": "Test User", "email": "new@test.com", "password": "Pass123!", "role": "patient"}', 
     'No user exists with email "new@test.com"'],
    
    ["TC-WB-SC-002", "Statement Coverage", "Admin Management", "pre('save') middleware in Admin.js", 
     "Create/update admin with password field modified", 
     "Pre-save hook checking `this.isModified('password')`, bcryptjs.hash() call with salt rounds=12, assignment `this.password = hashedPassword`", 
     'Password stored in DB is hashed (not plaintext), hash starts with "$2a$12$" (bcryptjs format)', 
     "POST", "http://localhost:5000/api/auth/admin/login (after admin creation)", 
     '{"email": "admin@healthconnect.com", "password": "#1ap@NITK"}', 
     "Admin created via createTestAdmin.js script"],
    
    ["TC-WB-SC-003", "Statement Coverage", "Authentication", "login() in authController.js", 
     "POST request with credentials for a suspended user (isActive=false)", 
     "User lookup, email verification check, statement `if (!user.isActive)`, logAuthEvent with failureReason='Account suspended', response with status 403 and suspensionReason", 
     'Status 403, {"success": false, "message": "Account suspended", "suspended": true, "suspensionReason": "Your account has been suspended..."}', 
     "POST", "http://localhost:5000/api/auth/login", 
     '{"email": "suspended@test.com", "password": "Pass123!", "role": "doctor"}', 
     'Doctor with email "suspended@test.com" exists, isEmailVerified=true, isActive=false'],
    
    ["TC-WB-SC-004", "Statement Coverage", "Authentication", "login() in authController.js", 
     "Valid credentials, but multiple recent OTP requests already sent (rate limit exceeded)", 
     "Credential validation passes, statement `OTP.checkRateLimit(email, 'login')`, condition `if (!rateLimitCheck.allowed)`, response with status 429 and waitTime", 
     'Status 429, {"success": false, "message": "Please wait X seconds before requesting a new OTP", "waitTime": X}', 
     "POST", "http://localhost:5000/api/auth/login", 
     '{"email": "patient@test.com", "password": "CorrectPass123!", "role": "patient"}', 
     "Patient exists with correct credentials, multiple OTP records exist within rate limit window"],
    
    ["TC-WB-SC-005", "Statement Coverage", "Authentication", "verifyOTP() for login in authController.js", 
     "Valid OTP for login purpose", 
     "OTP verification succeeds, purpose=='login' branch, statements: `previousLoginTime = user.lastLogin`, `user.lastLogin = new Date()`, `await user.save()`, `await generateToken()`, response includes loginInfo object", 
     'User lastLogin field updated in DB, response includes {"loginInfo": {"previousLogin": <timestamp>, "lastLogout": <timestamp>}}', 
     "POST", "http://localhost:5000/api/auth/verify-otp", 
     '{"email": "doctor@test.com", "otp": "123456", "role": "doctor", "purpose": "login"}', 
     'Valid OTP exists for email "doctor@test.com" with purpose "login", doctor exists and isEmailVerified=true']
]

# B. Branch Coverage
bc_data = [
    ["TC-WB-BC-001", "Branch Coverage", "Authentication", "register() in authController.js", 
     "POST request with email that already exists in Doctor or Patient collection", 
     "Lookup for existingDoctor and existingPatient, branch `if (existingDoctor || existingPatient)` evaluates to TRUE, logAuthEvent with failureReason='Email already registered', early return with 400 status", 
     'Status 400, {"success": false, "message": "User already exists with this email"}', 
     "POST", "http://localhost:5000/api/auth/register", 
     '{"name": "Duplicate User", "email": "existing@test.com", "password": "Pass123!", "role": "patient"}', 
     'Patient or Doctor with email "existing@test.com" already exists'],
    
    ["TC-WB-BC-002", "Branch Coverage", "Admin Authentication", "comparePassword() method in Admin.js, called from adminLogin()", 
     "Two scenarios: (1) Correct password (branch TRUE), (2) Incorrect password (branch FALSE)", 
     "Method `comparePassword(candidatepassword)`, bcryptjs.compare() returns boolean, if TRUE: authentication succeeds, if FALSE: returns 401 error", 
     'Scenario 1 (correct): Status 200, login successful; Scenario 2 (incorrect): Status 401, "Invalid credentials"', 
     "POST", "http://localhost:5000/api/auth/admin/login", 
     'Scenario 1: {"email": "admin@healthconnect.com", "password": "#1ap@NITK"} / Scenario 2: {"email": "admin@healthconnect.com", "password": "WrongPass"}', 
     'Admin exists with email "admin@healthconnect.com" and correct hashed password'],
    
    ["TC-WB-BC-003", "Branch Coverage", "OTP Verification", "verifyOTP() in authController.js", 
     'Two scenarios: (1) purpose="registration", (2) purpose="login"', 
     "OTP verified successfully, branch `if (purpose === 'registration')`: sets isEmailVerified=true, sends welcome email, returns user data; branch `if (purpose === 'login')`: updates lastLogin, generates JWT token, returns login info", 
     'Scenario 1 (registration): Email verified, welcome email sent, message "Email verified successfully! You can now login."; Scenario 2 (login): JWT cookie set, message "Login successful", loginInfo included', 
     "POST", "http://localhost:5000/api/auth/verify-otp", 
     'Scenario 1: {"email": "new@test.com", "otp": "123456", "role": "patient", "purpose": "registration"} / Scenario 2: {"email": "verified@test.com", "otp": "654321", "role": "doctor", "purpose": "login"}', 
     "Scenario 1: Valid registration OTP exists, user isEmailVerified=false; Scenario 2: Valid login OTP exists, user isEmailVerified=true"],
    
    ["TC-WB-BC-004", "Branch Coverage", "Admin Authentication", "adminLogin() in authController.js", 
     "Two scenarios: (1) isActive=true (branch TRUE), (2) isActive=false (branch FALSE)", 
     "Admin found, branch `if (!admin.isActive)`: if TRUE (inactive), return 403 with suspension message; if FALSE (active), proceed to password verification", 
     'Scenario 1 (active): Proceeds to password check; Scenario 2 (inactive): Status 403, {"success": false, "message": "Account is suspended. Please contact support."}', 
     "POST", "http://localhost:5000/api/auth/admin/login", 
     '{"email": "admin@healthconnect.com", "password": "#1ap@NITK"}', 
     "Scenario 1: Admin isActive=true; Scenario 2: Admin isActive=false"],
    
    ["TC-WB-BC-005", "Branch Coverage", "Authentication", "login() in authController.js", 
     "Two scenarios: (1) isEmailVerified=true (branch FALSE, proceeds), (2) isEmailVerified=false (branch TRUE, blocks login)", 
     "User found, branch `if (!user.isEmailVerified)`: if TRUE, log failed login, return 400 with requiresVerification; if FALSE, proceed to isActive check", 
     'Scenario 1 (verified): Proceeds to account status check; Scenario 2 (not verified): Status 400, {"success": false, "message": "Email not verified. Please verify your email first.", "requiresVerification": true}', 
     "POST", "http://localhost:5000/api/auth/login", 
     '{"email": "test@test.com", "password": "Pass123!", "role": "patient"}', 
     "Scenario 1: Patient isEmailVerified=true; Scenario 2: Patient isEmailVerified=false"]
]

# C. Multiple Condition Coverage
mc_data = [
    ["TC-WB-MC-001", "Multiple Condition Coverage", "Authentication", "register() in authController.js", 
     "Valid registration data, but email service (sendOTPEmail) fails", 
     "User created successfully, OTP generated, sendOTPEmail() called and returns {success: false}, condition `if (!emailResult.success)` evaluates TRUE, returns 201 with warning message and includes OTP in development mode", 
     'Status 201, {"success": true, "message": "Registration successful. Email service temporarily unavailable. Please contact support for OTP.", "data": {"email": "test@test.com", "requiresVerification": true, "otp": "123456"}} (in dev mode)', 
     "POST", "http://localhost:5000/api/auth/register", 
     '{"name": "Test User", "email": "test@test.com", "password": "Pass123!", "role": "patient"}', 
     "No existing user, email service configured to fail (or SMTP unavailable)"],
    
    ["TC-WB-MC-002", "Multiple Condition Coverage", "Password Reset", "forgotPassword() in authController.js", 
     "User who has already used password reset once (passwordResetCount >= 1)", 
     "User found, conditions `if (user.passwordResetCount && user.passwordResetCount >= 1)` both TRUE, logAuthEvent with failureReason='Password reset limit exceeded', return 403 with resetLimitReached=true", 
     'Status 403, {"success": false, "message": "Your 1 attempt for password reset has been completed. You cannot reset your password again...", "resetLimitReached": true, "resetCount": 1, "lastResetDate": <timestamp>}', 
     "POST", "http://localhost:5000/api/auth/forgot-password", 
     '{"email": "limited@test.com", "role": "patient"}', 
     "Patient exists with passwordResetCount=1"],
    
    ["TC-WB-MC-003", "Multiple Condition Coverage", "Password Reset", "resetPassword() in authController.js", 
     "OTP that is both expired (expiresAt < now) AND already verified (verified=true)", 
     "OTP lookup with conditions email, otp, purpose, AND verified=false; query returns null because verified=true, condition `if (!otpDoc)` TRUE, log failed attempt, return 400 \"Invalid or already used OTP\"", 
     'Status 400, {"success": false, "message": "Invalid, expired, or already used OTP. Please request a new password reset."}', 
     "POST", "http://localhost:5000/api/auth/reset-password", 
     '{"email": "test@test.com", "otp": "123456", "password": "NewPass123!", "confirmPassword": "NewPass123!", "role": "patient"}', 
     "OTP exists with verified=true (already used) OR is expired"],
    
    ["TC-WB-MC-004", "Multiple Condition Coverage", "Password Change", "changePassword() in authController.js", 
     "Authenticated user tries to change password to same current password", 
     "Current password verified (isPasswordCorrect=true), then `isSamePassword = await user.comparePassword(newPassword)`, condition `if (isSamePassword)` TRUE, return 400 \"New password must be different\"", 
     'Status 400, {"success": false, "message": "New password must be different from current password"}', 
     "POST", "http://localhost:5000/api/auth/change-password", 
     '{"currentPassword": "MyPass123!", "newPassword": "MyPass123!"}', 
     'Authenticated patient/doctor with password "MyPass123!", valid JWT token in cookie'],
    
    ["TC-WB-MC-005", "Multiple Condition Coverage", "Password Reset", "resetPassword() in authController.js", 
     "Valid OTP but password and confirmPassword fields do not match", 
     "Condition `if (password !== confirmPassword)` evaluates TRUE, immediate return with 400 status before OTP lookup", 
     'Status 400, {"success": false, "message": "Passwords do not match"}', 
     "POST", "http://localhost:5000/api/auth/reset-password", 
     '{"email": "test@test.com", "otp": "123456", "password": "NewPass123!", "confirmPassword": "DifferentPass123!", "role": "patient"}', 
     "Valid OTP exists (not checked due to early return)"]
]

# D. Path Coverage
pc_data = [
    ["TC-WB-PC-001", "Path Coverage", "Authentication (Full Flow)", 
     "register() → verifyOTP() (registration) → login() → verifyOTP() (login)", 
     "Step 1: Register new patient; Step 2: Verify registration OTP; Step 3: Login; Step 4: Verify login OTP", 
     "Complete path: User creation → isEmailVerified=false → OTP sent → OTP verification → isEmailVerified=true → welcome email → login request → credentials valid → login OTP sent → login OTP verified → JWT token issued → lastLogin updated", 
     'Final: Status 200, {"success": true, "message": "Login successful", user data, JWT cookie set}', 
     "POST (multiple calls)", "/api/auth/register → /api/auth/verify-otp → /api/auth/login → /api/auth/verify-otp", 
     "Multi-step (see description)", 
     "No existing user, clean state"],
    
    ["TC-WB-PC-002", "Path Coverage", "Authentication", "login() in authController.js", 
     "User exists with correct password but isEmailVerified=false", 
     "Path: User found → Password correct (not checked yet) → isEmailVerified check FAILS → logAuthEvent with failureReason → return 400 requiresVerification", 
     'Status 400, {"success": false, "message": "Email not verified. Please verify your email first.", "requiresVerification": true}', 
     "POST", "http://localhost:5000/api/auth/login", 
     '{"email": "unverified@test.com", "password": "Pass123!", "role": "patient"}', 
     'Patient with email "unverified@test.com", isEmailVerified=false, password set'],
    
    ["TC-WB-PC-003", "Path Coverage", "Admin Authentication", "adminLogin() → generateToken() → Session.create()", 
     "Valid admin credentials", 
     "Path: Admin found → isActive=true → Password verified → previousLoginTime captured → lastlogin updated → generateToken called → JWT created → Session record created in DB → Cookie set → loginInfo prepared → Response sent", 
     'Status 200, admin data returned, JWT cookie "token" set, Session record in DB with isActive=true', 
     "POST", "http://localhost:5000/api/auth/admin/login", 
     '{"email": "admin@healthconnect.com", "password": "#1ap@NITK"}', 
     "Admin exists, isActive=true"],
    
    ["TC-WB-PC-004", "Path Coverage", "Password Reset", "forgotPassword() → resetPassword() → sendPasswordChangedEmail()", 
     "Step 1: Request password reset; Step 2: Submit valid OTP and new password", 
     "Path: User found → resetCount < 1 → OTP generated → Email sent → OTP verified → Passwords match → OTP marked verified → Password hashed → passwordResetCount incremented → passwordResetUsedAt set → User saved → Password changed email sent → Success response", 
     'Status 200, {"success": true, "message": "Password reset successful! You can now login with your new password."}, password changed email sent', 
     "POST (two calls)", "/api/auth/forgot-password → /api/auth/reset-password", 
     'Step 1: {"email": "user@test.com", "role": "patient"} / Step 2: {"email": "user@test.com", "otp": "123456", "password": "NewPass!", "confirmPassword": "NewPass!", "role": "patient"}', 
     "Patient exists with passwordResetCount=0 or undefined"],
    
    ["TC-WB-PC-005", "Path Coverage", "Authentication", "logout() in authController.js", 
     "Authenticated user with valid JWT token in cookie", 
     "Path: Token extracted from cookie → userId and userRole identified → UserModel.findByIdAndUpdate sets lastLogout → Session.updateOne sets isActive=false → Cookie cleared (two attempts: secure=true and secure=false) → Success response", 
     'Status 200, {"success": true, "message": "Logged out successfully"}, Session isActive=false in DB, lastLogout timestamp updated, cookie "token" cleared', 
     "POST", "http://localhost:5000/api/auth/logout", 
     "None (uses cookie)", 
     "User authenticated with active session, valid JWT token in cookie"]
]

# E. Data Flow-Based Testing
df_data = [
    ["TC-WB-DF-001", "Data Flow-Based Testing", "Admin Model", "Admin.create() → pre('save') → comparePassword()", 
     "Create admin with plaintext password, then login with same password", 
     "DU chain: password variable defined at Admin.create({password: \"plaintext\"}) → DEFINITION in pre-save hook: `this.password = hashedPassword` → USE in comparePassword: `bcryptjs.compare(candidatepassword, this.password)`", 
     "Password hashed correctly, comparePassword returns true for correct plaintext, login succeeds", 
     "POST", "http://localhost:5000/api/auth/admin/login", 
     '{"email": "admin@healthconnect.com", "password": "#1ap@NITK"}', 
     "Admin created with password field"],
    
    ["TC-WB-DF-002", "Data Flow-Based Testing", "OTP Management", "OTP.createOTP() → OTP.verifyOTP() → OTP.deleteOne()", 
     "Request OTP for login, then verify it", 
     "DU chain: otp variable DEFINED in createOTP: `otp = Math.floor(100000 + Math.random() * 900000).toString()` → USED in verifyOTP: `OTP.findOne({email, otp, purpose})` → USED in delete: `await OTP.deleteOne({ _id: otpDoc._id })` after successful verification", 
     "OTP created in DB, verified successfully, then deleted after use; cannot be reused", 
     "POST (two calls)", "/api/auth/login → /api/auth/verify-otp", 
     'Step 1: {"email": "doctor@test.com", "password": "Pass123!", "role": "doctor"} / Step 2: {"email": "doctor@test.com", "otp": "<generated>", "role": "doctor", "purpose": "login"}', 
     "Doctor exists with verified email"],
    
    ["TC-WB-DF-003", "Data Flow-Based Testing", "Authentication", "register() in authController.js", 
     'Register with role="doctor"', 
     "DU chain: role variable DEFINED from req.body → USED in `getUserModel(role)` to select Doctor vs Patient model → USED in `formatUserResponse(user, role)` to include role-specific fields (specialization, experience) → USED in response JSON", 
     'Doctor model used, registration successful, response includes doctor-specific fields if role="doctor"', 
     "POST", "http://localhost:5000/api/auth/register", 
     '{"name": "Dr. Smith", "email": "drsmith@hospital.com", "password": "Pass123!", "role": "doctor", "specialization": "Cardiology", "experience": 5, "qualification": "MD"}', 
     "No existing user with this email"],
    
    ["TC-WB-DF-004", "Data Flow-Based Testing", "OTP Resend", "resendOTP() in authController.js", 
     "Request OTP resend for existing user", 
     "DU chain: email variable DEFINED from req.body → USED in `UserModel.findOne({ email })` → USED in `OTP.checkRateLimit(email, purpose)` → USED in `OTP.createOTP(email, role, purpose)` → USED in `sendOTPEmail({ email: user.email, ... })`", 
     "Email used consistently throughout flow, OTP sent to correct email address", 
     "POST", "http://localhost:5000/api/auth/resend-otp", 
     '{"email": "patient@test.com", "role": "patient", "purpose": "registration"}', 
     'Patient exists with email "patient@test.com", not rate limited'],
    
    ["TC-WB-DF-005", "Data Flow-Based Testing", "Session Management", "generateToken() → logout()", 
     "Login successfully, then logout", 
     "DU chain: token DEFINED in generateToken: `const token = jwt.sign(...)` → USED in Session.create({token, ...}) → USED in res.cookie('token', token) → USED in logout: `const { token } = req.cookies` → USED in `Session.updateOne({ token, isActive: true }, { isActive: false })` → Cookie cleared", 
     "Token created, stored in DB session, set as cookie, successfully revoked on logout, cookie cleared", 
     "POST (login then logout)", "/api/auth/verify-otp (login) → /api/auth/logout", 
     "Login OTP verification, then logout (no body)", 
     "Successful login creates session with token"]
]


def add_data_to_sheet(sheet, headers, data_sections, section_titles):
    """Add data to sheet with headers for each section and blank row separation"""
    current_row = 1
    
    for idx, (data, title) in enumerate(zip(data_sections, section_titles)):
        # Add section title
        sheet.cell(row=current_row, column=1, value=title)
        sheet.cell(row=current_row, column=1).font = subtype_font
        sheet.cell(row=current_row, column=1).fill = subtype_fill
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        current_row += 1
        
        # Add headers
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        current_row += 1
        
        # Add data rows
        for row_data in data:
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=current_row, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = border
            current_row += 1
        
        # Add blank row after each section (except last)
        if idx < len(data_sections) - 1:
            current_row += 1

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        if header in ["Test_Case_ID", "Technique", "Request_Method"]:
            sheet.column_dimensions[column_letter].width = 20
        elif header in ["Module", "Function", "Input_Field"]:
            sheet.column_dimensions[column_letter].width = 25
        elif header in ["Boundary_Checked"]:
            sheet.column_dimensions[column_letter].width = 30
        else:
            sheet.column_dimensions[column_letter].width = 50


# Add Black-Box data
bb_sections = [ecp_data, bva_data]
bb_titles = [
    "A. Equivalence Class Partitioning (5 Test Cases)",
    "B. Boundary Value Analysis (5 Test Cases)"
]
add_data_to_sheet(bb_sheet, bb_headers, bb_sections, bb_titles)

# Add White-Box data
wb_sections = [sc_data, bc_data, mc_data, pc_data, df_data]
wb_titles = [
    "A. Statement Coverage (5 Test Cases)",
    "B. Branch Coverage (5 Test Cases)",
    "C. Multiple Condition Coverage (5 Test Cases)",
    "D. Path Coverage (5 Test Cases)",
    "E. Data Flow-Based Testing (5 Test Cases)"
]
add_data_to_sheet(wb_sheet, wb_headers, wb_sections, wb_titles)

# Save workbook
output_file = "Health_Connect_Test_Cases.xlsx"
wb.save(output_file)
print(f"✅ Excel file created successfully: {output_file}")
print(f"📊 Total test cases: 35 (10 Black-Box + 25 White-Box)")
print(f"📁 Black-Box subtypes: 2 (Equivalence Class Partitioning, Boundary Value Analysis)")
print(f"📁 White-Box subtypes: 5 (Statement, Branch, Multiple Condition, Path, Data Flow)")
print(f"\n✨ Each subtype has:")
print(f"   - Section title row (with background color)")
print(f"   - Header row (with column names)")
print(f"   - Data rows (5 test cases per subtype)")
print(f"   - Blank row separator (between subtypes)")
